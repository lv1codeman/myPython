# 載入需要的套件
import numpy as np
from util import get_num_column_dict
from util import is_contain_chinese
from util import getSyllabusColumns

from datetime import datetime
import openpyxl
import pandas as pd
import os

# import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import re
import csv
import sys
from selenium.webdriver.chrome.options import Options
import logging


def search_courses(year=112, semester=2, crsid="", crsclassID="", crossclass=""):
    # 開啟瀏覽器視窗(Chrome)
    # 方法一：執行前需開啟chromedriver.exe且與執行檔在同一個工作目錄
    print("start Chrome")
    logging.getLogger("selenium").setLevel(logging.WARNING)
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_experimental_option("excludeSwitches", ["disable-logging"])
    driver = webdriver.Chrome(options=chrome_options)
    driver.get("http://webapt.ncue.edu.tw/DEANV2/Other/ob010")
    print("END Chrome")

    # 指定條件區的輸入
    # 指定學年度<select>
    select_element = driver.find_element(By.ID, "ddl_yms_year")
    select = Select(select_element)
    select.select_by_value(year)
    # 指定學期<select>
    select_element = driver.find_element(By.ID, "ddl_yms_smester")
    select = Select(select_element)
    select.select_by_value(semester)

    # 指定課程代碼<input>
    if len(crsid) == 5:
        select_element = driver.find_element(By.ID, "scr_selcode")
        select_element.send_keys(crsid)

    # 指定修課班別<select>
    select = Select(driver.find_element(By.ID, "ddl_scj_cls_id"))
    select.select_by_value(crsclassID)

    # 點擊「查詢」按鈕
    button = driver.find_element(By.XPATH, "//input[@value='查詢']")
    button.click()

    delay = 10  # seconds

    try:
        # 直到開課課程列表的序號出現
        myElem = WebDriverWait(driver, delay).until(
            EC.presence_of_element_located((By.CLASS_NAME, "tb_01"))
        )
        print("Page is ready!")
        # 畫面截圖（檢查用）
        driver.save_screenshot("python_test.png")

        html = driver.page_source
        driver.close()  # 關閉瀏覽器
        soup = BeautifulSoup(html, "html.parser")

        results = soup.find(id="result")
        results_row = results.select("tbody tr")

        # 搜尋到的資料筆數 = row的大小
        row = len(results_row)
        # 建立放output的陣列
        output = [[0] * 20 for i in range(row)]

        results = results.select("tbody td")

        i = 0
        deli = 0
        unit = ""  # 開課單位

        for result in results:
            if result.get("data-th") == "課程代碼：":
                output[i][0] = result.get_text()
            elif result.get("data-th") == "開課班別(代表)：":
                output[i][1] = result.get_text()
                unit = result.get_text()
            elif result.get("data-th") == "課程名稱：":
                course_name_all = result.text.split("\n")
                course_name_list = list()
                for ele in course_name_all:
                    if len(ele.strip()) >= 2:
                        course_name_list.append(ele.strip())
                if len(course_name_list) > 1:
                    output[i][2] = course_name_list[0]
                    output[i][3] = course_name_list[1]
                else:
                    output[i][2] = course_name_list[0]
                    output[i][3] = ""
            elif result.get("data-th") == "教學大綱：":
                hasCHT = not "無檔案" in result.text
                hasENG = not "No file" in result.text
                res = getSyllabusColumns(result, hasCHT, hasENG)
                output[i][4] = res[0]
                output[i][5] = res[1]
                output[i][6] = res[2]
            elif result.get("data-th") == "課程性質：":
                output[i][7] = result.get_text()
            elif result.get("data-th") == "課程性質2：":
                output[i][8] = result.get_text()
            elif result.get("data-th") == "全英語授課：":
                output[i][9] = result.get_text()
            elif result.get("data-th") == "學分：":
                output[i][10] = result.get_text()
            elif result.get("data-th") == "教師姓名：":
                teacher_list = list()
                teachers = result.find_all("span")
                for teacher in teachers:
                    teacher_list.append(teacher.get_text())
                teachers = ",".join(str(element) for element in teacher_list)
                output[i][11] = teachers.strip()
            elif result.get("data-th") == "上課大樓：":
                output[i][12] = result.get_text()
            elif result.get("data-th") == "上課節次+地點：":
                output[i][13] = result.get_text()
            elif result.get("data-th") == "上限人數：":
                output[i][14] = result.get_text()
            elif result.get("data-th") == "登記人數：":
                output[i][15] = re.sub(r"\s+", "", result.get_text())
            elif result.get("data-th") == "選上人數：":
                output[i][16] = result.get_text()
                if int(result.get_text()) < 10:
                    if ("碩" in unit) & (int(result.get_text()) >= 3):
                        output[i][17] = "Y"
                    elif (
                        ("博" in unit)
                        & ("碩" not in unit)
                        & (int(result.get_text()) >= 1)
                    ):
                        output[i][17] = "Y"
                    else:
                        output[i][17] = "N"
                else:
                    output[i][17] = "Y"
            elif result.get("data-th") == "可跨班：":
                match crossclass:
                    case "":
                        deli = 1
                        output[i][18] = result.get_text()
                    case "可跨班系":
                        if result.get_text() == "可跨班系":
                            deli = 1
                            output[i][18] = result.get_text()
                        else:
                            deli = 0
                            output[i][18] = result.get_text()
                    case "限本系":
                        if result.get_text() == "限本系":
                            deli = 1
                            output[i][18] = result.get_text()
                        else:
                            deli = 0
                            output[i][18] = result.get_text()
                    case "限本班":
                        if result.get_text() == "限本班":
                            deli = 1
                            output[i][18] = result.get_text()
                        else:
                            deli = 0
                            output[i][18] = result.get_text()
            elif result.get("data-th") == "備註：":
                output[i][19] = result.get_text().strip()
                if deli == 1:
                    i = i + 1
                else:
                    for k in range(0, 19):
                        output[i][k] = 0

    except TimeoutException:
        print("Loading took too much time!")

    # 首欄不為0者才存入output等待輸出
    output = [row for row in output if row[0] != 0]
    print("Total output:", len(output), "rows.")

    file_name = year + semester + "_course_list.csv"
    # 開啟輸出的 CSV 檔案(準備寫入檔案)
    with open(file_name, "w", newline="", encoding="utf-8-sig") as csvfile:
        # 建立 CSV 檔寫入器
        writer = csv.writer(csvfile)

        # 寫入一列資料
        writer.writerow(
            [
                "課程代碼",
                "開課班別",
                "課程名稱(中)",
                "課程名稱(英)",
                "教學大綱(中)",
                "教學大綱(英)",
                "未填教學大綱",
                "課程性質",
                "課程性質2",
                "全英語授課",
                "學分",
                "教師姓名",
                "上課大樓",
                "上課節次+地點",
                "上限人數",
                "登記人數",
                "選上人數",
                "符合開課標準",
                "可跨班",
                "備註",
            ]
        )

        # 寫入另外幾列資料
        for i, length in enumerate(output):
            writer.writerow(
                [
                    output[i][0],
                    output[i][1],
                    output[i][2],
                    output[i][3],
                    output[i][4],
                    output[i][5],
                    output[i][6],
                    output[i][7],
                    output[i][8],
                    output[i][9],
                    output[i][10],
                    output[i][11],
                    output[i][12],
                    output[i][13],
                    output[i][14],
                    output[i][15],
                    output[i][16],
                    output[i][17],
                    output[i][18],
                    output[i][19],
                ]
            )
        # 關閉檔案
        csvfile.close()

    # 做vlookup查出系所和承辦人
    # 讀入表1
    # df1 = pd.read_csv("系所對照表.csv", encoding="UTF-8-sig")
    # # 讀入表2
    # df2 = pd.read_csv(file_name, encoding="UTF-8-sig")
    # # 關聯數據
    # data = df2.merge(
    #     df1, on="開課班別", left_index=False, right_index=False, sort=False, how="left"
    # )
    # # 保存數據
    # data.to_csv(file_name, encoding="utf-8-sig", index=False)

    csvfile = open(file_name, encoding="utf-8-sig")  # 開啟 CSV 檔案
    raw_data = csv.reader(csvfile)  # 讀取 CSV 檔案
    data = list(raw_data)  # 轉換成二維串列

    wb = openpyxl.Workbook()  # 建立空白的 Excel 活頁簿物件
    # sheet = wb.create_sheet("csv")  # 建立空白的工作表
    s1 = wb["Sheet"]
    for i in data:
        s1.append(i)  # 逐筆添加到最後一列

    s1.title = year + "-" + semester + "開課查詢"
    xlsx_filename = (
        year
        + "-"
        + semester
        + "開課查詢_"
        + datetime.now().strftime("%Y-%m-%d")
        + ".xlsx"
    )

    max_column = s1.max_column
    max_column_dict = {}
    column_width = 0
    for i in range(1, max_column + 1):
        sheet_value_list = [k for k in str(s1.cell(row=1, column=i).value)]
        for v in sheet_value_list:
            if is_contain_chinese(v) == True:
                column_width += 2.2
            else:
                column_width += 0.8
        max_column_dict[i] = column_width + 0.5
        column_width = 0

    num_str_dict = get_num_column_dict()
    # print(num_str_dict)
    for key, value in max_column_dict.items():
        s1.column_dimensions[num_str_dict[key]].width = value

    wb.save(xlsx_filename)
    # 關閉並移除csv檔
    csvfile.close()
    os.remove(file_name)

    # df = pd.read_excel('112-2開課查詢_2024-03-08.xlsx')

    # print(df)

    print("done")

    return 0


if __name__ == "__main__":
    while True:
        try:
            print("請輸入搜尋條件或輸入q離開...")

            year = input("學年度: ")
            if year == "q":
                sys.exit(0)
            if re.match(r"^(9[5-9]|1[0-4][0-9]|150)$", year) is None:
                raise ValueError("[必填欄位] 學年度錯誤(例: 112)")

            semester = input("學期: ")
            if semester == "q":
                sys.exit(0)
            if re.match(r"^[1234]{1}$", semester) is None:
                raise ValueError("[必填欄位] 學期錯誤(例: 1)")

            crsid = input("課程代碼(為空則查詢全部): ")
            if crsid == "q":
                sys.exit(0)
            if re.match(r"^$|([A-Z0-9]{2}\d{3})$", crsid) is None:
                raise ValueError("[非必填欄位] 課程代碼僅限英數字")
        except ValueError as err:
            print("輸入有誤，請檢查輸入的值")
            print("錯誤訊息: ", err.args)
            os.system("pause")
            continue
        else:
            break

    search_courses(year, semester, crsid)
