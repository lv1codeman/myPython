# 載入需要的套件
from util import is_contain_chinese
from util import get_num_column_dict
from datetime import datetime
import openpyxl
import pandas as pd
import os
# import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import re
import csv

# 開啟瀏覽器視窗(Chrome)
# 方法一：執行前需開啟chromedriver.exe且與執行檔在同一個工作目錄
driver = webdriver.Chrome()
# 方法二：或是直接指定exe檔案路徑
# driver = webdriver.Chrome('./chromedriver')
driver.get("http://webapt.ncue.edu.tw/DEANV2/Other/ob010")  # 更改網址以前往不同網頁

# 改搜尋資料改這裡
year = "112"  # 學年度
semester = "2"  # 學期

# 定位搜尋框
# <select>
select_element = driver.find_element(By.ID, "ddl_yms_year")
select = Select(select_element)
select.select_by_value(year)

select_element = driver.find_element(By.ID, "ddl_yms_smester")
select = Select(select_element)
select.select_by_value(semester)

# 查詢指定課程代碼
# <input>
# select_element = driver.find_element(By.ID, "scr_selcode")
# select_element.send_keys("71041")

# 查詢指定修課班別
# select_element = driver.find_element(By.ID, "ddl_scj_cls_id")
# select = Select(select_element)
# select.select_by_value('D110BN1A')

# 點擊「查詢」按鈕
button = driver.find_element(By.XPATH, "//input[@value='查詢']")
button.click()

delay = 10  # seconds

try:
    # 直到開課課程列表的序號出現
    myElem = WebDriverWait(driver, delay).until(
        EC.presence_of_element_located((By.CLASS_NAME, "tb_01"))
    )
    print("Page is ready!")
    # 畫面截圖（檢查用）
    driver.save_screenshot("python_test.png")
    html = driver.page_source
    driver.close()  # 關閉瀏覽器
    soup = BeautifulSoup(html, "html.parser")
    results = soup.find(id="result")
    results_row = results.select("tbody tr")
    # 搜尋到的資料筆數 = row的大小
    row = len(results_row)
    # 建立放output的陣列
    output = [[0] * 20 for i in range(row)]
    results = results.select("tbody td")
    i = 0
    unit = ""  # 開課單位
    for result in results:
        if result.get("data-th") == "課程代碼：":
            output[i][0] = result.get_text()
        elif result.get("data-th") == "開課班別(代表)：":
            output[i][1] = result.get_text()
            unit = result.get_text()
        elif result.get("data-th") == "課程名稱：":
            course_name_all = result.text.split("\n")
            course_name_list = list()
            for ele in course_name_all:
                if len(ele.strip()) >= 2:
                    course_name_list.append(ele.strip())
            if len(course_name_list) > 1:
                output[i][2] = course_name_list[0]
                output[i][3] = course_name_list[1]
            else:
                output[i][2] = course_name_list[0]
                output[i][3] = ""
        elif result.get("data-th") == "教學大綱：":
            if ("無檔案" in result.text) & ("No file" in result.text):
                output[i][4] = "無檔案"
                output[i][5] = "No file"
                output[i][6] = "Y"
            else:
                output[i][6] = "N"
                if ("無檔案" in result.text) & (not "No file" in result.text):
                    cht_syllabus = "無檔案"
                    eng_syllabus = result.find("a").text.strip()
                    output[i][4] = cht_syllabus
                    output[i][5] = eng_syllabus
                elif not ("No file" in result.text) & (not "無檔案" in result.text):
                    cht_syllabus = result.find("a").text.strip()
                    eng_syllabus = "No file"
                    output[i][4] = cht_syllabus
                    output[i][5] = eng_syllabus
                else:
                    cht_syllabus = result.find("a").text.strip()
                    eng_syllabus = result.find(
                        "a"
                    ).next_sibling.next_sibling.next_sibling.strip()
                    output[i][4] = cht_syllabus
                    output[i][5] = eng_syllabus
        elif result.get("data-th") == "課程性質：":
            output[i][7] = result.get_text()
        elif result.get("data-th") == "課程性質2：":
            output[i][8] = result.get_text()
        elif result.get("data-th") == "全英語授課：":
            output[i][9] = result.get_text()
        elif result.get("data-th") == "學分：":
            output[i][10] = result.get_text()
        elif result.get("data-th") == "教師姓名：":
            teacher_list = list()
            teachers = result.find_all("span")
            for teacher in teachers:
                teacher_list.append(teacher.get_text())
            teachers = ",".join(str(element) for element in teacher_list)
            output[i][11] = teachers.strip()
        elif result.get("data-th") == "上課大樓：":
            output[i][12] = result.get_text()
        elif result.get("data-th") == "上課節次+地點：":
            output[i][13] = result.get_text()
        elif result.get("data-th") == "上限人數：":
            output[i][14] = result.get_text()
        elif result.get("data-th") == "登記人數：":
            output[i][15] = re.sub(r"\s+", "", result.get_text())
        elif result.get("data-th") == "選上人數：":
            output[i][16] = result.get_text()
            if int(result.get_text()) < 10:
                if ("碩" in unit) & (int(result.get_text()) >= 3):
                    output[i][19] = "Y"
                elif (
                    ("博" in unit) & ("碩" not in unit) & (
                        int(result.get_text()) >= 1)
                ):
                    output[i][19] = "Y"
                else:
                    output[i][19] = "N"
            else:
                output[i][19] = "Y"
        elif result.get("data-th") == "可跨班：":
            output[i][17] = result.get_text()
        elif result.get("data-th") == "備註：":
            output[i][18] = result.get_text().strip()
            i = i + 1
except TimeoutException:
    print("Loading took too much time!")

file_name = year + semester + "_course_list.csv"
# 開啟輸出的 CSV 檔案(準備寫入檔案)
with open(file_name, "w", newline="", encoding="utf-8-sig") as csvfile:
    # 建立 CSV 檔寫入器
    writer = csv.writer(csvfile)

    # 寫入一列資料
    writer.writerow(
        [
            "課程代碼",
            "開課班別(代表)",
            "課程名稱(中)",
            "課程名稱(英)",
            "教學大綱(中)",
            "教學大綱(英)",
            "未填教學大綱",
            "課程性質",
            "課程性質2",
            "全英語授課",
            "學分",
            "教師姓名",
            "上課大樓",
            "上課節次+地點",
            "上限人數",
            "登記人數",
            "選上人數",
            "可跨班",
            "備註",
            "符合開課標準",
        ]
    )

    # 寫入另外幾列資料
    for i, length in enumerate(output):
        writer.writerow(
            [
                output[i][0],
                output[i][1],
                output[i][2],
                output[i][3],
                output[i][4],
                output[i][5],
                output[i][6],
                output[i][7],
                output[i][8],
                output[i][9],
                output[i][10],
                output[i][11],
                output[i][12],
                output[i][13],
                output[i][14],
                output[i][15],
                output[i][16],
                output[i][17],
                output[i][18],
                output[i][19],
            ]
        )
    # 關閉檔案
    csvfile.close()
csvfile = open(file_name, encoding="utf-8-sig")  # 開啟 CSV 檔案
raw_data = csv.reader(csvfile)  # 讀取 CSV 檔案
data = list(raw_data)  # 轉換成二維串列

wb = openpyxl.Workbook()  # 建立空白的 Excel 活頁簿物件
s1 = wb["Sheet"]
for i in data:
    s1.append(i)  # 逐筆添加到最後一列

s1.title = year + "-" + semester + "開課查詢"
xlsx_filename = "開課查詢_" + datetime.now().strftime("%Y-%m-%d") + ".xlsx"

max_column = s1.max_column
max_column_dict = {}
column_width = 0

for i in range(1, max_column+1):
    sheet_value_list = [k for k in str(s1.cell(row=1, column=i).value)]
    for v in sheet_value_list:
        if is_contain_chinese(v) == True:
            column_width += 2.2
        else:
            column_width += 0.8
    max_column_dict[i] = column_width + 0.5
    column_width = 0

num_str_dict = get_num_column_dict()
for key, value in max_column_dict.items():
    s1.column_dimensions[num_str_dict[key]].width = value
wb.save(xlsx_filename)
# 關閉並移除csv檔
csvfile.close()
os.remove(file_name)
print("done")
